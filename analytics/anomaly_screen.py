#!/usr/bin/env python3
"""
C-Drishti · Cooperative procurement anomaly screen
Public portfolio edition

Reproduces the 3-sigma screening method used in the C-Drishti dashboard.
The public repository ships with a synthetic demonstration dataset so the
analytics workflow is reproducible without redistributing restricted source data.

Usage:
    pip install -r analytics/requirements.txt
    python analytics/anomaly_screen.py data/sample_procurement_data.xlsx
    python analytics/anomaly_screen.py data/sample_procurement_data.xlsx --chart out.png

Method (a screening rule, not an accusation):
  metric  = pending lift as % of quantity procured, per centre
  flag if = metric > mean + 3*sigma (population sd), computed across the file

Limitations:
  * A heavily right-skewed distribution can make a 3-sigma rule misleading.
    Production systems should compare centres against peer, district and seasonal
    baselines and should validate results with domain experts.
  * High pending stock can reflect legitimate logistics constraints. A flag is an
    audit-priority signal only; it is not evidence of fraud or wrongdoing.
"""

import argparse
import statistics as st
from pathlib import Path

import openpyxl


def load(path: str, min_qtl: float) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows: list[dict] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue

        try:
            procured = float(row[4] or 0)
            pending = float(row[10] or 0)
        except (TypeError, ValueError):
            continue

        if procured > min_qtl:
            rows.append(
                {
                    "district": str(row[1]).strip(),
                    "society": str(row[2]).strip(),
                    "centre": str(row[3]).strip(),
                    "procured": procured,
                    "pending": pending,
                    "pct": pending / procured * 100.0,
                }
            )

    return rows


def analyse(rows: list[dict]) -> tuple[float, float, float, list[dict]]:
    if not rows:
        raise ValueError("No eligible rows were found in the supplied workbook.")

    values = [row["pct"] for row in rows]
    mean = st.mean(values)
    sd = st.pstdev(values)
    threshold = mean + 3 * sd

    for row in rows:
        row["z"] = (row["pct"] - mean) / sd if sd else 0.0

    flagged = sorted(
        (row for row in rows if row["pct"] > threshold),
        key=lambda row: -row["pct"],
    )
    return mean, sd, threshold, flagged


def save_chart(flagged: list[dict], mean: float, total: int, top_n: int, path: str) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    top = flagged[:top_n][::-1]
    if not top:
        print("\nNo flagged rows; chart was not generated.")
        return

    figure, axis = plt.subplots(figsize=(9, 0.5 * len(top) + 1.5), dpi=150)
    axis.barh(range(len(top)), [row["pct"] for row in top])
    axis.axvline(mean, linestyle="--", linewidth=1.2)
    axis.set_yticks(range(len(top)))
    axis.set_yticklabels([row["centre"][:34] for row in top], fontsize=9)
    axis.set_xlabel("pending lift as % of quantity procured")
    axis.set_title(f"3-sigma screening · n={total} · mean {mean:.2f}%")
    plt.tight_layout()
    plt.savefig(path)
    print(f"\nchart saved: {path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Run the C-Drishti 3-sigma screening workflow.")
    parser.add_argument("xlsx", help="path to a procurement reconciliation workbook")
    parser.add_argument(
        "--min-qtl",
        type=float,
        default=1000,
        help="ignore centres below this procurement volume (default: 1000)",
    )
    parser.add_argument("--top", type=int, default=10, help="number of flagged rows to print")
    parser.add_argument("--chart", help="optional PNG path for a bar chart")
    args = parser.parse_args()

    workbook_path = Path(args.xlsx)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    rows = load(str(workbook_path), args.min_qtl)
    mean, sd, threshold, flagged = analyse(rows)

    print(f"centres analysed : {len(rows)} (procurement > {args.min_qtl:g} qtl)")
    print(f"mean             : {mean:.2f}%")
    print(f"sigma            : {sd:.2f}")
    print(f"3-sigma line     : {threshold:.2f}%")
    print(f"flagged          : {len(flagged)} centres\n")
    print(f"{'#':>2}  {'pending%':>8}  {'z':>6}  {'stuck qtl':>10}  centre")

    for index, row in enumerate(flagged[: args.top], 1):
        print(
            f"{index:>2}  {row['pct']:>7.1f}%  {row['z']:>5.1f}σ  "
            f"{row['pending']:>10,.0f}  {row['centre'][:60]}"
        )

    if args.chart:
        save_chart(flagged, mean, len(rows), args.top, args.chart)


if __name__ == "__main__":
    main()
